import requests
from bs4 import BeautifulSoup
import openpyxl


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = "British_Airways"
print(excel.sheetnames)
sheet.append(['Name', 'Other Info', 'Overall_review'])

for i in range(1,364):
    html_tags = requests.get(f"https://www.airlinequality.com/airline-reviews/british-airways/page/{i}/").text
    soup = BeautifulSoup(html_tags, 'html.parser')
    review = soup.find_all('div', class_="body")
    for rev in review:
        name_date = rev.find_next('h3', class_="text_sub_header userStatusWrapper").text
        date_start_index = name_date.rfind("(")
        name = name_date[:date_start_index].strip()
        tags = rev.find_all('tr')
        tag_texts = []
        for tag in tags:
            tag_text = tag.text.strip()
            tag_texts.append(tag_text)  # Append tag text to the list
        # Join tag texts using newline separator
        all_tag_text = '\n'.join(tag_texts)
        filled_stars = rev.find_all('span', class_='star fill')
        num_filled_stars = len(filled_stars)
        print("\n")
        sheet.append([name, all_tag_text, num_filled_stars])
    excel.save('British_Airways.xlsx')